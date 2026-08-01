#!/usr/bin/env python3
import os
import sys
import json
import uuid
import argparse
import datetime
import openpyxl

def main():
    parser = argparse.ArgumentParser(description="Generate Rapor Snapshot for ARSC Leaderboard")
    parser.add_argument("--excel", required=True, help="Path to Master Rapor Excel (for NIM)")
    parser.add_argument("--payload", required=True, help="Path to Rapor Payload JSON")
    parser.add_argument("--release-code", required=True, help="Release Code (e.g., RTP_2026)")
    args = parser.parse_args()

    # Load Excel to map Name -> NIM
    print(f"Loading Excel workbook from {args.excel}...")
    try:
        wb = openpyxl.load_workbook(args.excel, data_only=True)
    except Exception as e:
        print(f"Failed to load excel: {e}")
        sys.exit(1)
        
    if 'Data Anggota' not in wb.sheetnames:
        print("Sheet 'Data Anggota' not found in Excel!")
        sys.exit(1)
    
    ws = wb['Data Anggota']
    
    # Headers are on row 3
    # 2: Nama, 3: NIM
    name_to_nim = {}
    for r in range(4, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        nim = ws.cell(row=r, column=3).value
        if name and nim:
            name_to_nim[str(name).strip().lower()] = str(nim).strip()

    # Load or initialize identity map (NIM -> UUID)
    data_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), '.data')
    os.makedirs(data_dir, exist_ok=True)
    identity_map_path = os.path.join(data_dir, 'identity_map.json')
    
    identity_map = {}
    if os.path.exists(identity_map_path):
        with open(identity_map_path, 'r', encoding='utf-8') as f:
            identity_map = json.load(f)

    # Load payload
    print(f"Loading payload from {args.payload}...")
    with open(args.payload, 'r', encoding='utf-8') as f:
        payload_data = json.load(f)
        
    members = payload_data.get('members', [])
    
    minimized_snapshot = {
        "schema_version": "1.0",
        "generated_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "release_code": args.release_code,
        "members": []
    }
    
    synthetic_snapshot = {
        "schema_version": "1.0",
        "generated_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "release_code": args.release_code,
        "members": []
    }
    
    mapped_count = 0
    new_uuid_count = 0
    
    for m in members:
        # Resolve identity
        name = m.get('name', '').strip()
        nim = name_to_nim.get(name.lower())
        
        if not nim:
            print(f"WARNING: Could not find NIM for member {name}. Skipping.")
            continue
            
        if nim not in identity_map:
            identity_map[nim] = str(uuid.uuid4())
            new_uuid_count += 1
            
        member_id = identity_map[nim]
        
        # Build public record
        record = {
            "member_id": member_id,
            "canonical_name": name,
            "unit": m.get('unit'),
            "position": m.get('jabatan'),
            "release_code": args.release_code,
            "release_member_code": m.get('member_code'),
            "evaluation_status": m.get('status_penilaian')
        }
        minimized_snapshot["members"].append(record)
        mapped_count += 1

    # Save identity map
    with open(identity_map_path, 'w', encoding='utf-8') as f:
        json.dump(identity_map, f, indent=2)
    
    # Save minimized snapshot
    minimized_path = os.path.join(data_dir, 'MINIMIZED_CONFIDENTIAL_SNAPSHOT.json')
    with open(minimized_path, 'w', encoding='utf-8') as f:
        json.dump(minimized_snapshot, f, indent=2)
        
    # Generate synthetic snapshot (2 records)
    synthetic_members = [
        {
            "member_id": str(uuid.uuid4()),
            "canonical_name": "John Doe",
            "unit": "IT",
            "position": "Staf Ahli",
            "release_code": args.release_code,
            "release_member_code": f"{args.release_code}_001",
            "evaluation_status": "Dinilai"
        },
        {
            "member_id": str(uuid.uuid4()),
            "canonical_name": "Jane Smith",
            "unit": "PR",
            "position": "Anggota",
            "release_code": args.release_code,
            "release_member_code": f"{args.release_code}_002",
            "evaluation_status": "Dinilai"
        }
    ]
    synthetic_snapshot["members"] = synthetic_members
    synthetic_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'tests', 'fixtures', 'synthetic_snapshot.json')
    os.makedirs(os.path.dirname(synthetic_path), exist_ok=True)
    with open(synthetic_path, 'w', encoding='utf-8') as f:
        json.dump(synthetic_snapshot, f, indent=2)
        
    print(f"Snapshot generation complete.")
    print(f"Mapped {mapped_count} members. Generated {new_uuid_count} new UUIDs.")
    print(f"Saved: {minimized_path}")
    print(f"Saved: {synthetic_path}")

if __name__ == '__main__':
    main()
